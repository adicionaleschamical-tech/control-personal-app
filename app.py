import streamlit as st
import gspread
import pandas as pd
import datetime
import random
import time
import json
from collections import defaultdict
import io
import plotly.express as px
import plotly.graph_objects as go

# Zona horaria Argentina (UTC-3)
tz_arg = datetime.timezone(datetime.timedelta(hours=-3))

# Intentar importar openpyxl para Excel formateado
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

st.set_page_config(
    page_title="Sistema de Gestión de Personal", 
    layout="wide",
    page_icon="👮‍♂️",
    initial_sidebar_state="expanded"
)

# ========== CSS MEJORADO CON DISEÑO MODERNO ==========
st.markdown("""
<style>
    /* Importar fuente moderna */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
    
    * {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
    }
    
    /* Reset de márgenes y estilos base */
    html, body, [class*="css"] {
        font-size: 14px;
        line-height: 1.6;
    }
    
    /* ========== TÍTULOS ========== */
    h1 {
        font-size: 2.5rem !important;
        font-weight: 800 !important;
        background: linear-gradient(135deg, #1f3a6b 0%, #2c5a8c 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        margin-bottom: 0.5rem !important;
        letter-spacing: -0.5px;
    }
    
    h2 {
        font-size: 1.8rem !important;
        font-weight: 700 !important;
        color: #1f3a6b;
        border-left: 4px solid #2ecc71;
        padding-left: 15px;
        margin: 25px 0 20px 0 !important;
    }
    
    h3 {
        font-size: 1.3rem !important;
        font-weight: 600 !important;
        color: #2c5a8c;
        margin: 15px 0 !important;
    }
    
    /* ========== SIDEBAR MEJORADO ========== */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0f2b3d 0%, #1a3a4f 100%);
        box-shadow: 2px 0 15px rgba(0,0,0,0.15);
        padding-top: 20px;
    }
    
    [data-testid="stSidebar"] .stMarkdown, 
    [data-testid="stSidebar"] .stWrite {
        color: #e8f0f7 !important;
    }
    
    [data-testid="stSidebar"] .stButton button {
        background: linear-gradient(135deg, #2ecc71 0%, #27ae60 100%);
        color: white;
        border: none;
        border-radius: 10px;
        font-weight: 600;
        padding: 10px 15px;
        transition: all 0.3s ease;
        box-shadow: 0 2px 8px rgba(46,204,113,0.3);
    }
    
    [data-testid="stSidebar"] .stButton button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 15px rgba(46,204,113,0.5);
    }
    
    [data-testid="stSidebar"] .stButton button:last-child {
        background: linear-gradient(135deg, #e74c3c 0%, #c0392b 100%) !important;
        box-shadow: 0 2px 8px rgba(231,76,60,0.3);
    }
    
    [data-testid="stSidebar"] .stButton button:last-child:hover {
        box-shadow: 0 4px 15px rgba(231,76,60,0.5);
    }
    
    /* ========== MÉTRICAS EN TARJETAS ========== */
    div[data-testid="stMetric"] {
        background: white;
        border-radius: 15px;
        padding: 20px 15px;
        box-shadow: 0 4px 20px rgba(0,0,0,0.06);
        border: 1px solid #eef2f7;
        transition: all 0.3s ease;
    }
    
    div[data-testid="stMetric"]:hover {
        transform: translateY(-5px);
        box-shadow: 0 8px 30px rgba(0,0,0,0.12);
        border-color: #2ecc71;
    }
    
    div[data-testid="stMetric"] label {
        font-weight: 600 !important;
        font-size: 0.9rem !important;
        color: #2c5a8c !important;
    }
    
    div[data-testid="stMetric"] div[data-testid="stMetricValue"] {
        font-size: 2.2rem !important;
        font-weight: 700 !important;
        color: #1f3a6b !important;
    }
    
    /* ========== TABLAS ========== */
    .stDataFrame {
        border-radius: 12px;
        overflow: hidden;
        box-shadow: 0 2px 12px rgba(0,0,0,0.06);
        border: 1px solid #eef2f7;
    }
    
    .stDataFrame div[data-testid="stElement"]:hover {
        box-shadow: 0 4px 20px rgba(0,0,0,0.1);
    }
    
    /* ========== BOTONES ========== */
    .stButton button {
        border-radius: 10px;
        font-weight: 600;
        padding: 10px 20px;
        transition: all 0.3s ease;
        background: linear-gradient(135deg, #3498db 0%, #2980b9 100%);
        color: white;
        border: none;
        box-shadow: 0 2px 8px rgba(52,152,219,0.3);
    }
    
    .stButton button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 15px rgba(52,152,219,0.4);
    }
    
    .stButton button[data-testid="baseButton-secondary"] {
        background: linear-gradient(135deg, #95a5a6 0%, #7f8c8d 100%);
        box-shadow: 0 2px 8px rgba(149,165,166,0.3);
    }
    
    /* ========== ALERTAS Y MENSAJES ========== */
    .stAlert {
        border-radius: 12px;
        border-left: 4px solid;
        padding: 15px 20px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.04);
    }
    
    .stAlert .stMarkdown {
        font-weight: 500;
    }
    
    /* ========== EXPANDERS ========== */
    .streamlit-expanderHeader {
        font-weight: 600;
        color: #1f3a6b;
        font-size: 0.95rem;
        padding: 10px 15px;
        background-color: #f8fafc;
        border-radius: 8px;
        margin: 2px 0;
        transition: background-color 0.2s;
        border: 1px solid #eef2f7;
    }
    
    .streamlit-expanderHeader:hover {
        background-color: #e8edf4;
    }
    
    /* ========== FOOTER ========== */
    .footer {
        position: fixed;
        bottom: 0;
        left: 0;
        right: 0;
        background: linear-gradient(135deg, #1f3a6b 0%, #2c5a8c 100%);
        color: white;
        text-align: center;
        padding: 12px;
        z-index: 999;
        font-size: 0.75rem;
        box-shadow: 0 -2px 10px rgba(0,0,0,0.1);
    }
    
    /* ========== CONTENEDORES ========== */
    .custom-container {
        background: white;
        border-radius: 15px;
        padding: 20px;
        margin: 15px 0;
        box-shadow: 0 2px 12px rgba(0,0,0,0.05);
        border: 1px solid #eef2f7;
    }
    
    /* ========== SELECTORES Y INPUTS ========== */
    .stSelectbox label, .stMultiSelect label, .stTextInput label {
        font-weight: 600 !important;
        font-size: 0.9rem !important;
        color: #1f3a6b !important;
    }
    
    /* ========== SEPARADORES ========== */
    hr {
        margin: 30px 0 !important;
        border: none !important;
        border-top: 2px solid #eef2f7 !important;
    }
    
    /* ========== OCULTAR ELEMENTOS POR DEFECTO ========== */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    
    /* ========== TABS PERSONALIZADOS ========== */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        background: #f8fafc;
        padding: 8px;
        border-radius: 12px;
    }
    
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px;
        padding: 8px 20px;
        font-weight: 600;
        font-size: 0.9rem;
        transition: all 0.3s ease;
    }
    
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #1f3a6b 0%, #2c5a8c 100%);
        color: white;
        box-shadow: 0 2px 8px rgba(31,58,107,0.3);
    }
    
    /* ========== FICHA DEL AGENTE ========== */
    .agente-ficha {
        background: linear-gradient(135deg, #ffffff 0%, #f8fafc 100%);
        border-radius: 20px;
        padding: 30px;
        box-shadow: 0 8px 32px rgba(0,0,0,0.08);
        border: 1px solid #e8edf4;
        margin: 15px 0;
        transition: all 0.3s ease;
    }
    
    .agente-ficha:hover {
        box-shadow: 0 12px 48px rgba(0,0,0,0.12);
        transform: translateY(-2px);
    }
    
    .agente-header {
        display: flex;
        align-items: center;
        gap: 20px;
        margin-bottom: 25px;
        padding-bottom: 20px;
        border-bottom: 2px solid #eef2f7;
    }
    
    .agente-avatar {
        width: 80px;
        height: 80px;
        border-radius: 50%;
        background: linear-gradient(135deg, #1f3a6b 0%, #2c5a8c 100%);
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 2.5rem;
        color: white;
        font-weight: 700;
        flex-shrink: 0;
        box-shadow: 0 4px 12px rgba(31,58,107,0.3);
    }
    
    .agente-nombre {
        font-size: 1.8rem;
        font-weight: 700;
        color: #1f3a6b;
        margin: 0;
    }
    
    .agente-badge {
        display: inline-block;
        padding: 4px 14px;
        border-radius: 20px;
        font-size: 0.8rem;
        font-weight: 600;
        margin-top: 6px;
    }
    
    .badge-jerarquia {
        background: #dbeafe;
        color: #1e40af;
    }
    
    .badge-funcion {
        background: #d1fae5;
        color: #065f46;
    }
    
    .badge-dependencia {
        background: #fef3c7;
        color: #92400e;
    }
    
    .agente-grid {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
        gap: 16px;
        margin-top: 15px;
    }
    
    .agente-campo {
        background: white;
        padding: 14px 18px;
        border-radius: 12px;
        border: 1px solid #eef2f7;
        transition: all 0.2s ease;
    }
    
    .agente-campo:hover {
        border-color: #2c5a8c;
        background: #fafcff;
    }
    
    .agente-campo-label {
        font-size: 0.7rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        color: #94a3b8;
        margin-bottom: 4px;
    }
    
    .agente-campo-valor {
        font-size: 1rem;
        font-weight: 500;
        color: #1e293b;
    }
    
    .agente-campo-valor.vacio {
        color: #94a3b8;
        font-style: italic;
    }
</style>
""", unsafe_allow_html=True)

# ------------------------------------------------------------
# Funciones de carga de datos (incluyendo Auditoria)
# ------------------------------------------------------------
@st.cache_data(ttl=3600)
def cargar_datos_hoja():
    with st.spinner("🔄 Cargando datos desde Google Sheets..."):
        creds = st.secrets["gsheets"]
        gc = gspread.service_account_from_dict(creds)
        sh = gc.open_by_key("1TH32e7TkB4RYEKxxRhGRnnYxjkT9AF9TeuplxtYih1Y")
        
        # ---------- Personal ----------
        ws_personal = sh.worksheet("Personal")
        all_values = ws_personal.get_all_values()
        if len(all_values) == 0:
            st.error("La hoja 'Personal' está vacía")
            st.stop()
        
        header = all_values[0]
        data = all_values[1:]
        if header:
            first_header = str(header[0]).strip().lower()
            if first_header == '' or first_header.isdigit() or first_header in ['n', 'num', 'numero', 'legajo', 'id']:
                header = header[1:]
                data = [row[1:] for row in data]
        df_personal = pd.DataFrame(data, columns=header)
        
        # ---------- Usuarios ----------
        ws_usuarios = sh.worksheet("Usuarios")
        all_users = ws_usuarios.get_all_values()
        if len(all_users) == 0:
            st.error("La hoja 'Usuarios' está vacía")
            st.stop()
        header_users = all_users[0]
        data_users = all_users[1:]
        if header_users:
            first_header_u = str(header_users[0]).strip().lower()
            if first_header_u == '' or first_header_u.isdigit() or first_header_u in ['n', 'num', 'numero', 'legajo', 'id']:
                header_users = header_users[1:]
                data_users = [row[1:] for row in data_users]
        df_usuarios = pd.DataFrame(data_users, columns=header_users)
        
        # ---------- Propuestas ----------
        try:
            ws_propuestas = sh.worksheet("Propuestas")
            propuestas_data = ws_propuestas.get_all_values()
            if len(propuestas_data) > 1:
                header_prop = propuestas_data[0]
                header_prop = [str(col).upper().strip() for col in header_prop]
                prop_data = propuestas_data[1:]
                df_propuestas = pd.DataFrame(prop_data, columns=header_prop)
            else:
                df_propuestas = pd.DataFrame(columns=['ID', 'FECHA', 'USUARIO_DNI', 'USUARIO_NOMBRE', 'DEPENDENCIA', 'ACCION', 'DATOS_ORIGINALES', 'DATOS_NUEVOS', 'ESTADO'])
        except gspread.exceptions.WorksheetNotFound:
            df_propuestas = pd.DataFrame(columns=['ID', 'FECHA', 'USUARIO_DNI', 'USUARIO_NOMBRE', 'DEPENDENCIA', 'ACCION', 'DATOS_ORIGINALES', 'DATOS_NUEVOS', 'ESTADO'])
            ws_propuestas = sh.add_worksheet(title="Propuestas", rows=1000, cols=20)
            ws_propuestas.update([list(df_propuestas.columns)])
        
        # ---------- Auditoria ----------
        try:
            ws_auditoria = sh.worksheet("Auditoria")
            auditoria_data = ws_auditoria.get_all_values()
            if len(auditoria_data) > 1:
                header_aud = auditoria_data[0]
                aud_data = auditoria_data[1:]
                df_auditoria = pd.DataFrame(aud_data, columns=header_aud)
            else:
                df_auditoria = pd.DataFrame(columns=['ID', 'FECHA', 'USUARIO_DNI', 'USUARIO_NOMBRE', 'DEPENDENCIA', 'ACCION', 'DETALLE'])
        except gspread.exceptions.WorksheetNotFound:
            df_auditoria = pd.DataFrame(columns=['ID', 'FECHA', 'USUARIO_DNI', 'USUARIO_NOMBRE', 'DEPENDENCIA', 'ACCION', 'DETALLE'])
            ws_auditoria = sh.add_worksheet(title="Auditoria", rows=1000, cols=20)
            ws_auditoria.update([list(df_auditoria.columns)])
        
        # Limpieza general
        for df in [df_personal, df_usuarios]:
            df.columns = df.columns.str.strip()
            for col in df.columns:
                df[col] = df[col].astype(str).str.strip()
            df.replace('', pd.NA, inplace=True)
        
        if 'JERARQUÍA' in df_personal.columns:
            df_personal['JERARQUÍA'] = df_personal['JERARQUÍA'].str.replace(r'\s+', ' ', regex=True).str.upper()
        if 'DEPENDENCIA' in df_personal.columns:
            df_personal['DEPENDENCIA'] = df_personal['DEPENDENCIA'].str.replace(r'\s+', ' ', regex=True).str.upper()
        
        return df_personal, df_usuarios, df_propuestas, df_auditoria, sh

# ------------------------------------------------------------
# Funciones de escritura
# ------------------------------------------------------------
def get_new_connection():
    creds = st.secrets["gsheets"]
    gc = gspread.service_account_from_dict(creds)
    return gc.open_by_key("1TH32e7TkB4RYEKxxRhGRnnYxjkT9AF9TeuplxtYih1Y")

def registrar_auditoria(usuario_dni, usuario_nombre, dependencia, accion, detalle, sh):
    try:
        ws = sh.worksheet("Auditoria")
        auditoria_data = ws.get_all_values()
        if len(auditoria_data) > 1:
            header = auditoria_data[0]
            aud_data = auditoria_data[1:]
            df_aud = pd.DataFrame(aud_data, columns=header)
        else:
            df_aud = pd.DataFrame(columns=['ID', 'FECHA', 'USUARIO_DNI', 'USUARIO_NOMBRE', 'DEPENDENCIA', 'ACCION', 'DETALLE'])
        
        nuevo_id = len(df_aud) + 1 if not df_aud.empty else 1
        fecha_arg = datetime.datetime.now(tz_arg).strftime("%Y-%m-%d %H:%M:%S")
        
        nueva_fila = pd.DataFrame([{
            'ID': nuevo_id,
            'FECHA': fecha_arg,
            'USUARIO_DNI': usuario_dni,
            'USUARIO_NOMBRE': usuario_nombre,
            'DEPENDENCIA': dependencia,
            'ACCION': accion,
            'DETALLE': detalle
        }])
        
        df_aud = pd.concat([df_aud, nueva_fila], ignore_index=True)
        
        ws.clear()
        ws.update([df_aud.columns.tolist()] + df_aud.values.tolist())
        return True
    except Exception as e:
        st.error(f"Error al registrar auditoría: {e}")
        return False

def reordenar_por_jerarquia():
    try:
        sh = get_new_connection()
        ws = sh.worksheet("Personal")
        all_data = ws.get_all_values()
        header = all_data[0]
        data = all_data[1:]
        
        df = pd.DataFrame(data, columns=header)
        
        orden_jerarquia = {
            "COMISARIO MAYOR": 1,
            "COMISARIO INSPECTOR": 2,
            "COMISARIO": 3,
            "SUB COMISARIO": 4,
            "OFICIAL PRINCIPAL": 5,
            "OFICIAL INSPECTOR": 6,
            "OFICIAL SUB INSPECTOR": 7,
            "OFICIAL AYUDANTE": 8,
            "SUB OFICIAL MAYOR": 9,
            "SUB OFICIAL AUXILIAR": 10,
            "SUB OFICIAL ESCRIBIENTE": 11,
            "SARGENTO PRIMERO": 12,
            "SARGENTO": 13,
            "CABO PRIMERO": 14,
            "CABO": 15,
            "AGENTE": 16
        }
        
        df['_orden_jerarquia'] = df['JERARQUÍA'].map(orden_jerarquia).fillna(99)
        df = df.sort_values(['DEPENDENCIA', '_orden_jerarquia'])
        df = df.drop(columns=['_orden_jerarquia'])
        
        if 'ID' in df.columns:
            df['ID'] = [f"ID_{i+1}" for i in range(len(df))]
        
        ws.clear()
        ws.update([header] + df.values.tolist())
        return True
    except Exception as e:
        st.error(f"Error al reordenar: {e}")
        return False

def guardar_propuesta(usuario_dni, usuario_nombre, dependencia, accion, datos_originales, datos_nuevos):
    try:
        sh = get_new_connection()
        ws = sh.worksheet("Propuestas")
        
        propuestas_data = ws.get_all_values()
        if len(propuestas_data) > 1:
            header_prop = propuestas_data[0]
            prop_data = propuestas_data[1:]
            propuestas_df = pd.DataFrame(prop_data, columns=header_prop)
        else:
            propuestas_df = pd.DataFrame(columns=['ID', 'FECHA', 'USUARIO_DNI', 'USUARIO_NOMBRE', 'DEPENDENCIA', 'ACCION', 'DATOS_ORIGINALES', 'DATOS_NUEVOS', 'ESTADO'])
        
        nuevo_id = len(propuestas_df) + 1 if not propuestas_df.empty else 1
        fecha_arg = datetime.datetime.now(tz_arg).strftime("%Y-%m-%d %H:%M:%S")
        
        nueva_propuesta = pd.DataFrame([{
            'ID': nuevo_id,
            'FECHA': fecha_arg,
            'USUARIO_DNI': usuario_dni,
            'USUARIO_NOMBRE': usuario_nombre,
            'DEPENDENCIA': dependencia,
            'ACCION': accion,
            'DATOS_ORIGINALES': json.dumps(datos_originales, ensure_ascii=False),
            'DATOS_NUEVOS': json.dumps(datos_nuevos, ensure_ascii=False),
            'ESTADO': 'PENDIENTE'
        }])
        
        propuestas_df = pd.concat([propuestas_df, nueva_propuesta], ignore_index=True)
        
        ws.clear()
        ws.update([propuestas_df.columns.tolist()] + propuestas_df.values.tolist())
        
        st.session_state.df_propuestas = propuestas_df
        return True
    except Exception as e:
        st.error(f"Error al guardar propuesta: {e}")
        return False

def aprobar_propuesta(id_propuesta, user_dni, user_nombre):
    try:
        sh = get_new_connection()
        
        ws_propuestas = sh.worksheet("Propuestas")
        propuestas_data = ws_propuestas.get_all_values()
        if len(propuestas_data) <= 1:
            return False
        
        header_prop = propuestas_data[0]
        prop_data = propuestas_data[1:]
        propuestas_df = pd.DataFrame(prop_data, columns=header_prop)
        
        propuesta = propuestas_df[propuestas_df['ID'] == str(id_propuesta)].iloc[0]
        
        if propuesta['ESTADO'] != 'PENDIENTE':
            return False
        
        accion = propuesta['ACCION']
        datos_nuevos = json.loads(propuesta['DATOS_NUEVOS'])
        datos_originales = json.loads(propuesta['DATOS_ORIGINALES'])
        
        ws_personal = sh.worksheet("Personal")
        all_data = ws_personal.get_all_values()
        header = all_data[0]
        
        detalle = ""
        dependencia_afectada = propuesta['DEPENDENCIA']
        
        if accion == 'AGREGAR':
            ultimo_id = 0
            if 'ID' in header:
                col_id_idx = header.index('ID')
                for row in all_data[1:]:
                    if len(row) > col_id_idx and row[col_id_idx].startswith('ID_'):
                        try:
                            num = int(row[col_id_idx].split('_')[1])
                            ultimo_id = max(ultimo_id, num)
                        except:
                            pass
            
            nuevo_id = ultimo_id + 1
            datos_nuevos['ID'] = f"ID_{nuevo_id}"
            
            nueva_fila = [datos_nuevos.get(col, '') for col in header]
            ws_personal.append_row(nueva_fila)
            
            detalle = f"Agregado nuevo agente: {datos_nuevos.get('APELLIDO Y NOMBRE', '')} (DNI: {datos_nuevos.get('DNI', '')})"
        
        elif accion == 'MODIFICAR':
            dni_modificar = datos_nuevos.get('DNI')
            nombre_modificar = datos_nuevos.get('APELLIDO Y NOMBRE')
            
            if dni_modificar:
                col_dni_idx = header.index('DNI') if 'DNI' in header else None
                if col_dni_idx is not None:
                    for i, row in enumerate(all_data[1:], start=2):
                        if len(row) > col_dni_idx and row[col_dni_idx] == str(dni_modificar):
                            cambios = []
                            for col_idx, col_name in enumerate(header):
                                if col_name in datos_nuevos and col_name != 'ID':
                                    original = datos_originales.get(col_name, '')
                                    nuevo = datos_nuevos.get(col_name, '')
                                    if str(original) != str(nuevo):
                                        cambios.append(f"{col_name}: {original} → {nuevo}")
                                        ws_personal.update_cell(i, col_idx+1, str(nuevo))
                            detalle = f"Modificación de {nombre_modificar} (DNI: {dni_modificar}): " + "; ".join(cambios) if cambios else "Sin cambios detectados"
                            break
            elif nombre_modificar:
                col_nombre_idx = header.index('APELLIDO Y NOMBRE') if 'APELLIDO Y NOMBRE' in header else None
                if col_nombre_idx is not None:
                    for i, row in enumerate(all_data[1:], start=2):
                        if len(row) > col_nombre_idx and row[col_nombre_idx] == nombre_modificar:
                            cambios = []
                            for col_idx, col_name in enumerate(header):
                                if col_name in datos_nuevos and col_name != 'ID':
                                    original = datos_originales.get(col_name, '')
                                    nuevo = datos_nuevos.get(col_name, '')
                                    if str(original) != str(nuevo):
                                        cambios.append(f"{col_name}: {original} → {nuevo}")
                                        ws_personal.update_cell(i, col_idx+1, str(nuevo))
                            detalle = f"Modificación de {nombre_modificar}: " + "; ".join(cambios) if cambios else "Sin cambios detectados"
                            break
        
        elif accion == 'ELIMINAR':
            dni_eliminar = datos_nuevos.get('DNI')
            nombre_eliminar = datos_nuevos.get('APELLIDO Y NOMBRE')
            if dni_eliminar:
                col_dni_idx = header.index('DNI') if 'DNI' in header else None
                if col_dni_idx is not None:
                    for i, row in enumerate(all_data[1:], start=2):
                        if len(row) > col_dni_idx and row[col_dni_idx] == str(dni_eliminar):
                            detalle = f"Eliminado agente: {nombre_eliminar} (DNI: {dni_eliminar})"
                            ws_personal.delete_rows(i)
                            break
        
        # Registrar en auditoría
        if detalle:
            registrar_auditoria(user_dni, user_nombre, dependencia_afectada, accion, detalle, sh)
        
        # Reordenar por jerarquía
        reordenar_por_jerarquia()
        
        # Actualizar estado de la propuesta
        propuestas_df.loc[propuestas_df['ID'] == str(id_propuesta), 'ESTADO'] = 'APROBADO'
        ws_propuestas.clear()
        ws_propuestas.update([propuestas_df.columns.tolist()] + propuestas_df.values.tolist())
        
        # Recargar datos
        st.session_state.df_personal, st.session_state.df_usuarios, st.session_state.df_propuestas, st.session_state.df_auditoria, st.session_state.sh = cargar_datos_hoja()
        return True
    except Exception as e:
        st.error(f"Error al aprobar propuesta: {e}")
        return False

def rechazar_propuesta(id_propuesta):
    try:
        sh = get_new_connection()
        
        ws_propuestas = sh.worksheet("Propuestas")
        propuestas_data = ws_propuestas.get_all_values()
        if len(propuestas_data) <= 1:
            return False
        
        header_prop = propuestas_data[0]
        prop_data = propuestas_data[1:]
        propuestas_df = pd.DataFrame(prop_data, columns=header_prop)
        
        propuestas_df.loc[propuestas_df['ID'] == str(id_propuesta), 'ESTADO'] = 'RECHAZADO'
        
        ws_propuestas.clear()
        ws_propuestas.update([propuestas_df.columns.tolist()] + propuestas_df.values.tolist())
        
        st.session_state.df_propuestas = propuestas_df
        return True
    except Exception as e:
        st.error(f"Error al rechazar propuesta: {e}")
        return False

# Carga inicial
if 'df_personal' not in st.session_state:
    resultado = cargar_datos_hoja()
    st.session_state.df_personal = resultado[0]
    st.session_state.df_usuarios = resultado[1]
    st.session_state.df_propuestas = resultado[2]
    st.session_state.df_auditoria = resultado[3]
    st.session_state.sh = resultado[4]

if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False
if 'filtros' not in st.session_state:
    st.session_state.filtros = {}
if 'propuesta_rotacion' not in st.session_state:
    st.session_state.propuesta_rotacion = None
if 'generando_id' not in st.session_state:
    st.session_state.generando_id = False

# ==================== LOGIN ====================
if not st.session_state.logged_in:
    col1, col2 = st.columns([1, 1])
    with col1:
        st.title("👮‍♂️ Sistema de Gestión de Personal")
        st.markdown("### Iniciar Sesión")
        dni_input = st.text_input("📄 DNI", placeholder="Ingrese su número de documento")
        clave_input = st.text_input("🔒 Clave", type="password", placeholder="Ingrese su contraseña")
        if st.button("🚪 Ingresar", type="primary", use_container_width=True):
            if dni_input and clave_input:
                # Aquí debes adaptar la columna de la hoja "Usuarios"
                # Si tu hoja tiene "USUARIO" en lugar de "DNI", cambia la línea:
                # (st.session_state.df_usuarios['USUARIO'].astype(str).str.lower() == dni_input.lower())
                # y
                # (st.session_state.df_usuarios['CLAVE'].astype(str).str.lower() == clave_input.lower())
                usuario = st.session_state.df_usuarios[
                    (st.session_state.df_usuarios['DNI'].astype(str).str.lower() == dni_input.lower()) &
                    (st.session_state.df_usuarios['CLAVE'].astype(str).str.lower() == clave_input.lower())
                ]
                if not usuario.empty:
                    st.session_state.logged_in = True
                    st.session_state.user_data = usuario.iloc[0]
                    st.rerun()
                else:
                    st.error("❌ DNI o Clave incorrectos")
            else:
                st.warning("⚠️ Por favor, complete DNI y Clave")
    with col2:
        st.markdown("### ℹ️ Información")
        st.info("""
        **Sistema de Gestión de Personal**
        
        - 🔐 Acceso restringido a personal autorizado
        - 🔄 Gestión de rotaciones por jerarquía
        - 📎 Exportación de datos
        - 📊 Resúmenes estadísticos
        
        ---
        *Si no posee credenciales, contacte a **Personal de la Regional Quinta**.*
        """)
else:
    user = st.session_state.user_data
    with st.sidebar:
        st.markdown("### 👤 Panel de Usuario")
        st.markdown("---")
        st.markdown(f"**Nombre:** {user['NOMBRE']}")
        st.markdown(f"**📄 DNI:** {user['DNI']}")
        st.markdown(f"**🏢 Dependencia:** {user['DEPENDENCIA']}")
        st.markdown(f"**⭐ Jerarquía:** {user['JERARQUÍA']}")
        st.markdown(f"**📋 Función:** {user['FUNCIÓN']}")
        st.markdown("---")
        
        if st.button("🔄 REFRESCAR DATOS", use_container_width=True):
            with st.spinner("Actualizando datos..."):
                resultado = cargar_datos_hoja()
                st.session_state.df_personal = resultado[0]
                st.session_state.df_usuarios = resultado[1]
                st.session_state.df_propuestas = resultado[2]
                st.session_state.df_auditoria = resultado[3]
                st.session_state.sh = resultado[4]
                st.session_state.propuesta_rotacion = None
                st.success("✅ Datos actualizados correctamente")
                time.sleep(1)
                st.rerun()
        
        st.markdown("---")
        if st.button("🚪 Cerrar Sesión", use_container_width=True):
            st.session_state.logged_in = False
            st.session_state.propuesta_rotacion = None
            st.rerun()
    
    ahora_arg = datetime.datetime.now(tz_arg).strftime("%d/%m/%Y %H:%M:%S")
    st.title("👮‍♂️ Sistema de Gestión de Personal")
    st.caption(f"📅 Última actualización: {ahora_arg}")
    
    # ========== DETERMINAR ROL ==========
    rol = user['FUNCIÓN'].upper()
    es_admin = rol == "ADMINISTRADOR"
    es_supervisor = rol == "SUPERVISOR"
    es_usuario_comun = not es_admin and not es_supervisor
    
    # ========== PANEL DE AUDITORÍA (ÚLTIMOS 3 CAMBIOS) - VISIBLE PARA TODOS ==========
    st.markdown("## 📋 Últimos cambios registrados")
    if not st.session_state.df_auditoria.empty:
        df_aud = st.session_state.df_auditoria.copy()
        if 'FECHA' in df_aud.columns:
            df_aud['FECHA'] = pd.to_datetime(df_aud['FECHA'], errors='coerce')
            df_aud = df_aud.sort_values('FECHA', ascending=False).head(3)
            
            for _, row in df_aud.iterrows():
                fecha = row.get('FECHA', '')
                if isinstance(fecha, pd.Timestamp):
                    fecha = fecha.strftime("%d/%m/%Y %H:%M")
                usuario = row.get('USUARIO_NOMBRE', '')
                detalle = row.get('DETALLE', '')
                dependencia = row.get('DEPENDENCIA', '')
                accion = row.get('ACCION', '')
                icon = "📝"
                if accion == "AGREGAR":
                    icon = "➕"
                elif accion == "ELIMINAR":
                    icon = "➖"
                elif accion == "ROTACION":
                    icon = "🔄"
                st.info(f"{icon} **{fecha}** - {usuario} - {detalle} ({dependencia})")
        else:
            st.info("ℹ️ No hay registros de auditoría disponibles.")
    else:
        st.info("ℹ️ No hay registros de auditoría disponibles.")
    
    # ========== ALERTA DE PROPUESTAS PENDIENTES (SOLO ADMIN) ==========
    if es_admin:
        if st.button("🔄 Reordenar todo por Jerarquía", use_container_width=True):
            with st.spinner("Reordenando personal..."):
                if reordenar_por_jerarquia():
                    st.success("✅ Personal reordenado correctamente")
                    resultado = cargar_datos_hoja()
                    st.session_state.df_personal = resultado[0]
                    st.session_state.df_usuarios = resultado[1]
                    st.session_state.df_propuestas = resultado[2]
                    st.session_state.df_auditoria = resultado[3]
                    st.session_state.sh = resultado[4]
                    st.rerun()
                else:
                    st.error("❌ Error al reordenar")
        
        if not st.session_state.df_propuestas.empty:
            estado_col = None
            for col in st.session_state.df_propuestas.columns:
                if col.upper() == 'ESTADO':
                    estado_col = col
                    break
            
            if estado_col:
                propuestas_pendientes = len(st.session_state.df_propuestas[st.session_state.df_propuestas[estado_col].str.upper() == 'PENDIENTE'])
            else:
                propuestas_pendientes = 0
        else:
            propuestas_pendientes = 0
        
        if propuestas_pendientes > 0:
            st.warning(f"⚠️ **¡ATENCIÓN!** Hay {propuestas_pendientes} propuestas de cambio pendientes de revisar.", icon="⚠️")
            
            with st.expander(f"📋 Ver {propuestas_pendientes} propuestas pendientes", expanded=True):
                for idx, prop in st.session_state.df_propuestas.iterrows():
                    def get_col_value(row, posibles):
                        for p in posibles:
                            if p in row.index:
                                return row[p]
                        return "No disponible"
                    
                    prop_id = get_col_value(prop, ['ID', 'id'])
                    prop_fecha = get_col_value(prop, ['FECHA', 'Fecha'])
                    prop_usuario_nombre = get_col_value(prop, ['USUARIO_NOMBRE', 'Usuario_Nombre'])
                    prop_usuario_dni = get_col_value(prop, ['USUARIO_DNI', 'Usuario_DNI'])
                    prop_dependencia = get_col_value(prop, ['DEPENDENCIA', 'Dependencia'])
                    prop_accion = get_col_value(prop, ['ACCION', 'Accion'])
                    prop_estado = get_col_value(prop, ['ESTADO', 'Estado'])
                    
                    if prop_estado.upper() != 'PENDIENTE':
                        continue
                    
                    st.markdown(f"### 📋 Propuesta #{prop_id}")
                    st.markdown(f"**Fecha:** {prop_fecha}")
                    st.markdown(f"**Usuario:** {prop_usuario_nombre} (DNI: {prop_usuario_dni})")
                    st.markdown(f"**Dependencia:** {prop_dependencia}")
                    st.markdown(f"**Acción:** {prop_accion}")
                    
                    if prop_accion == 'MODIFICAR':
                        datos_originales_col = None
                        datos_nuevos_col = None
                        for col in prop.index:
                            if 'DATOS_ORIGINALES' in col.upper():
                                datos_originales_col = col
                            if 'DATOS_NUEVOS' in col.upper():
                                datos_nuevos_col = col
                        
                        if datos_originales_col and datos_nuevos_col:
                            try:
                                datos_originales = json.loads(prop[datos_originales_col])
                                datos_nuevos = json.loads(prop[datos_nuevos_col])
                                
                                st.markdown("#### 🔄 Cambios propuestos:")
                                cambios = []
                                todas_columnas = set(datos_originales.keys()) | set(datos_nuevos.keys())
                                for columna in todas_columnas:
                                    original = datos_originales.get(columna, '')
                                    nuevo = datos_nuevos.get(columna, '')
                                    if str(original) != str(nuevo):
                                        cambios.append({
                                            'Campo': columna,
                                            'Valor actual': original if original and str(original) != 'nan' else '(vacío)',
                                            'Valor propuesto': nuevo if nuevo and str(nuevo) != 'nan' else '(vacío)'
                                        })
                                
                                if cambios:
                                    st.table(pd.DataFrame(cambios))
                                else:
                                    st.info("No se detectaron cambios visibles.")
                                
                                with st.expander("Ver JSON completo"):
                                    st.json({"Original": datos_originales, "Nuevo": datos_nuevos})
                            except Exception as e:
                                st.error(f"Error al procesar datos: {e}")
                    
                    elif prop_accion == 'AGREGAR':
                        datos_nuevos_col = None
                        for col in prop.index:
                            if 'DATOS_NUEVOS' in col.upper():
                                datos_nuevos_col = col
                                break
                        if datos_nuevos_col:
                            try:
                                datos_nuevos = json.loads(prop[datos_nuevos_col])
                                st.markdown("#### ➕ Nuevo agente a agregar:")
                                st.json(datos_nuevos)
                            except:
                                st.write("Datos:", prop[datos_nuevos_col])
                    
                    elif prop_accion == 'ELIMINAR':
                        datos_originales_col = None
                        for col in prop.index:
                            if 'DATOS_ORIGINALES' in col.upper():
                                datos_originales_col = col
                                break
                        if datos_originales_col:
                            try:
                                datos_originales = json.loads(prop[datos_originales_col])
                                st.markdown("#### ➖ Agente a eliminar:")
                                st.json(datos_originales)
                            except:
                                st.write("Datos:", prop[datos_originales_col])
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button(f"✅ Aprobar #{prop_id}", key=f"aprobar_{prop_id}", use_container_width=True):
                            if aprobar_propuesta(prop_id, user['DNI'], user['NOMBRE']):
                                st.success(f"✅ Propuesta #{prop_id} aprobada")
                                st.rerun()
                            else:
                                st.error("❌ Error al aprobar")
                    with col2:
                        if st.button(f"❌ Rechazar #{prop_id}", key=f"rechazar_{prop_id}", use_container_width=True):
                            if rechazar_propuesta(prop_id):
                                st.success(f"❌ Propuesta #{prop_id} rechazada")
                                st.rerun()
                            else:
                                st.error("❌ Error al rechazar")
                    
                    st.markdown("---")
        else:
            st.info("✅ No hay propuestas de cambio pendientes.")
    
    # ========== CARGA DE DATOS SEGÚN ROL ==========
    if es_admin or es_supervisor:
        datos_completos = st.session_state.df_personal.copy()
        if es_admin:
            st.success("👑 **Modo Administrador** - Visualizando todo el personal", icon="👑")
        else:
            st.info("🔍 **Modo Supervisor** - Visualizando todo el personal (solo lectura)", icon="🔍")
        
        col_m1, col_m2, col_m3, col_m4 = st.columns(4)
        with col_m1:
            st.metric("📊 Total Personal", len(datos_completos))
        with col_m2:
            st.metric("🏢 Dependencias", datos_completos['DEPENDENCIA'].nunique() if 'DEPENDENCIA' in datos_completos else 0)
        with col_m3:
            st.metric("⭐ Jerarquías", datos_completos['JERARQUÍA'].nunique() if 'JERARQUÍA' in datos_completos else 0)
        with col_m4:
            st.metric("📋 Funciones", datos_completos['FUNCIÓN'].nunique() if 'FUNCIÓN' in datos_completos else 0)
    else:
        datos_completos = st.session_state.df_personal[st.session_state.df_personal['DEPENDENCIA'].str.lower() == user['DEPENDENCIA'].lower()].copy()
        st.info(f"👤 **Modo Usuario** - Visualizando personal de: {user['DEPENDENCIA']}", icon="ℹ️")
    
    if len(datos_completos) == 0:
        st.warning("⚠️ No hay personal para mostrar")
        st.stop()
    
    # ========== GENERAR ID ÚNICO ==========
    def generar_y_guardar_id():
        if st.session_state.generando_id:
            return False
        st.session_state.generando_id = True
        try:
            sh = get_new_connection()
            ws = sh.worksheet("Personal")
            all_data = ws.get_all_values()
            header = all_data[0]
            if 'ID' in header:
                return True
            new_header = header + ['ID']
            num_rows = len(all_data)
            new_data = []
            for i, row in enumerate(all_data):
                if i == 0:
                    new_row = row + ['ID']
                else:
                    new_row = row + [f"ID_{i}"]
                new_data.append(new_row)
            ws.update(range_name=f"A1:{chr(65+len(new_header)-1)}{num_rows}", values=new_data)
            resultado = cargar_datos_hoja()
            st.session_state.df_personal = resultado[0]
            st.session_state.df_usuarios = resultado[1]
            st.session_state.df_propuestas = resultado[2]
            st.session_state.df_auditoria = resultado[3]
            st.session_state.sh = resultado[4]
            return True
        except Exception as e:
            st.error(f"Error al generar ID: {e}")
            return False
        finally:
            st.session_state.generando_id = False
    
    tiene_dni = 'DNI' in datos_completos.columns
    tiene_id = 'ID' in datos_completos.columns
    identificador_col = None
    if tiene_dni:
        identificador_col = 'DNI'
    elif tiene_id:
        identificador_col = 'ID'
    else:
        identificador_col = None
        st.warning("⚠️ No se encontró una columna identificadora (DNI o ID).", icon="⚠️")
        if st.button("🔧 Generar ID único para todos los agentes", use_container_width=True):
            with st.spinner("Generando ID, espere un momento..."):
                exito = generar_y_guardar_id()
                if exito:
                    st.success("✅ ¡ID generado correctamente! La página se recargará.")
                    time.sleep(1)
                    st.rerun()
                else:
                    st.error("❌ No se pudo generar el ID. Revise la cuota de API o intente más tarde.")
        st.stop()
    
    # ========== DETECCIÓN DE COLUMNAS ==========
    def encontrar_columna(df, posibles):
        for p in posibles:
            if p in df.columns:
                return p
        return None
    
    nombre_col = encontrar_columna(datos_completos, ['APELLIDO Y NOMBRE', 'NOMBRE', 'NOMBRE COMPLETO', 'APELLIDO_NOMBRE'])
    jerarquia_col = encontrar_columna(datos_completos, ['JERARQUÍA', 'Jerarquia', 'JERARQUIA', 'RANGO', 'Rango', 'GRADO'])
    funcion_col = encontrar_columna(datos_completos, ['FUNCIÓN', 'Funcion', 'FUNCION', 'CARGO', 'Cargo', 'ROL'])
    dependencia_col = encontrar_columna(datos_completos, ['DEPENDENCIA', 'Dependencia', 'DEPARTAMENTO', 'UNIDAD'])
    sexo_col = encontrar_columna(datos_completos, ['SEXO', 'GENERO', 'GÉNERO', 'Sexo', 'Genero'])
    
    if not nombre_col or not jerarquia_col or not funcion_col or not dependencia_col:
        st.error("❌ Faltan columnas esenciales en la hoja de datos")
        st.stop()
    
    if sexo_col is None:
        st.info("ℹ️ No se detectó columna 'SEXO'. La rotación no tendrá en cuenta el género.")
    
    # ========== ORDEN JERÁRQUICO ==========
    orden_jerarquia_base = [
        "COMISARIO MAYOR", "COMISARIO INSPECTOR", "COMISARIO", "SUB COMISARIO",
        "OFICIAL PRINCIPAL", "OFICIAL INSPECTOR", "OFICIAL SUB INSPECTOR", "OFICIAL AYUDANTE",
        "SUB OFICIAL MAYOR", "SUB OFICIAL AUXILIAR", "SUB OFICIAL ESCRIBIENTE",
        "SARGENTO PRIMERO", "SARGENTO", "CABO PRIMERO", "CABO", "AGENTE"
    ]
    valores_existentes = datos_completos[jerarquia_col].unique()
    opciones_jerarquia = [j for j in orden_jerarquia_base if j in valores_existentes]
    opciones_jerarquia += sorted([j for j in valores_existentes if j not in orden_jerarquia_base])
    
    # ========== FILTROS DE BÚSQUEDA ==========
    st.markdown("## 🔎 Filtros de Búsqueda")
    col1, col2, col3 = st.columns(3)
    with col1:
        dep_filter = st.multiselect("🏢 Dependencia", sorted(datos_completos[dependencia_col].unique()))
    with col2:
        jer_filter = st.multiselect("⭐ Jerarquía", options=opciones_jerarquia)
    with col3:
        fun_filter = st.multiselect("📋 Función", sorted(datos_completos[funcion_col].unique()))
    
    with st.expander("➕ Filtros adicionales", expanded=False):
        columnas_extra = [col for col in datos_completos.columns 
                          if col not in [dependencia_col, jerarquia_col, funcion_col, identificador_col, nombre_col, sexo_col, 'SEXO', 'ID', 'DNI']]
        if columnas_extra:
            columna_extra = st.selectbox("Seleccione columna", ["Ninguna"] + columnas_extra)
            if columna_extra != "Ninguna":
                valores_extra = sorted(datos_completos[columna_extra].dropna().unique())
                valores_seleccionados = st.multiselect(f"Valores de {columna_extra}", valores_extra) if valores_extra else []
            else:
                valores_seleccionados = []
        else:
            columna_extra = "Ninguna"
            valores_seleccionados = []
    
    datos_filtrados = datos_completos.copy()
    if dep_filter:
        datos_filtrados = datos_filtrados[datos_filtrados[dependencia_col].isin(dep_filter)]
    if jer_filter:
        datos_filtrados = datos_filtrados[datos_filtrados[jerarquia_col].isin(jer_filter)]
    if fun_filter:
        datos_filtrados = datos_filtrados[datos_filtrados[funcion_col].isin(fun_filter)]
    if columna_extra != "Ninguna" and valores_seleccionados:
        datos_filtrados = datos_filtrados[datos_filtrados[columna_extra].isin(valores_seleccionados)]
    
    busqueda = st.text_input("🔍 Búsqueda rápida", placeholder="Nombre, DNI, dependencia...")
    if busqueda:
        mascara = datos_filtrados.astype(str).apply(lambda row: row.str.contains(busqueda, case=False).any(), axis=1)
        datos_filtrados = datos_filtrados[mascara]
        st.info(f"📌 {len(datos_filtrados)} resultados encontrados")
    
    # ========== GRÁFICOS AVANZADOS Y RESUMEN (SOLO ADMIN Y SUPERVISOR) ==========
    if es_admin or es_supervisor:
        st.markdown("## 📊 Análisis Visual de Personal")
        
        tab1, tab2, tab3, tab4 = st.tabs(["📊 Distribución", "⭐ Jerarquías", "📈 Ranking", "🔥 Mapa de Calor"])
        
        with tab1:
            col_chart1, col_chart2 = st.columns(2)
            
            with col_chart1:
                st.markdown("#### 🏆 Top 15 Dependencias")
                top_deps = datos_filtrados[dependencia_col].value_counts().head(15)
                if len(top_deps) > 0:
                    fig_bar = px.bar(
                        x=top_deps.values, 
                        y=top_deps.index,
                        orientation='h',
                        title="Cantidad de personal por dependencia",
                        labels={'x': 'Cantidad', 'y': 'Dependencia'},
                        color=top_deps.values,
                        color_continuous_scale='Blues'
                    )
                    fig_bar.update_layout(height=600, showlegend=False)
                    st.plotly_chart(fig_bar, use_container_width=True)
            
            with col_chart2:
                st.markdown("#### 🥧 Distribución porcentual")
                dep_counts = datos_filtrados[dependencia_col].value_counts()
                if len(dep_counts) > 0:
                    if len(dep_counts) > 15:
                        otros = dep_counts[15:].sum()
                        dep_counts = dep_counts[:15]
                        dep_counts['Otras dependencias'] = otros
                    fig_pie = px.pie(
                        values=dep_counts.values, 
                        names=dep_counts.index,
                        title="Porcentaje de personal por dependencia",
                        hole=0.3,
                        color_discrete_sequence=px.colors.qualitative.Set3
                    )
                    fig_pie.update_traces(textposition='inside', textinfo='percent+label')
                    fig_pie.update_layout(height=550)
                    st.plotly_chart(fig_pie, use_container_width=True)
        
        with tab2:
            col_jer1, col_jer2 = st.columns(2)
            
            with col_jer1:
                st.markdown("#### 📊 Distribución por Jerarquía")
                jer_counts = datos_filtrados[jerarquia_col].value_counts()
                if len(jer_counts) > 0:
                    fig_jer = px.bar(
                        x=jer_counts.values, 
                        y=jer_counts.index,
                        orientation='h',
                        title="Cantidad de personal por jerarquía",
                        labels={'x': 'Cantidad', 'y': 'Jerarquía'},
                        color=jer_counts.values,
                        color_continuous_scale='Greens'
                    )
                    fig_jer.update_layout(height=600, showlegend=False)
                    st.plotly_chart(fig_jer, use_container_width=True)
            
            with col_jer2:
                st.markdown("#### 🎯 Comparación por dependencia")
                if len(datos_filtrados[dependencia_col].unique()) > 0:
                    dep_seleccionada = st.selectbox("Seleccione dependencia:", datos_filtrados[dependencia_col].unique())
                    if dep_seleccionada:
                        dep_data = datos_filtrados[datos_filtrados[dependencia_col] == dep_seleccionada]
                        jer_dep = dep_data[jerarquia_col].value_counts()
                        if len(jer_dep) > 0:
                            fig_dep_jer = px.bar(
                                x=jer_dep.values, 
                                y=jer_dep.index,
                                orientation='h',
                                title=f"Jerarquías en {dep_seleccionada}",
                                labels={'x': 'Cantidad', 'y': 'Jerarquía'},
                                color=jer_dep.values,
                                color_continuous_scale='Oranges'
                            )
                            fig_dep_jer.update_layout(height=400)
                            st.plotly_chart(fig_dep_jer, use_container_width=True)
        
        with tab3:
            st.markdown("#### 🏆 Ranking de personal por dependencia")
            ranking_data = datos_filtrados[dependencia_col].value_counts().reset_index()
            ranking_data.columns = ['Dependencia', 'Cantidad']
            
            if len(ranking_data) > 0:
                max_count = ranking_data['Cantidad'].max()
                fig_ranking = go.Figure(data=[
                    go.Bar(
                        x=ranking_data['Cantidad'],
                        y=ranking_data['Dependencia'],
                        orientation='h',
                        marker_color=ranking_data['Cantidad'].apply(
                            lambda x: '#ef4444' if x > max_count*0.7 else 
                                     ('#f59e0b' if x > max_count*0.4 else '#10b981')
                        ),
                        text=ranking_data['Cantidad'],
                        textposition='outside'
                    )
                ])
                fig_ranking.update_layout(
                    title="Ranking de personal por dependencia",
                    xaxis_title="Cantidad de personal",
                    yaxis_title="Dependencia",
                    height=600,
                    showlegend=False
                )
                st.plotly_chart(fig_ranking, use_container_width=True)
                
                st.markdown("#### 📋 Tabla de ranking")
                ranking_data_sorted = ranking_data.sort_values('Cantidad', ascending=False)
                st.dataframe(ranking_data_sorted, use_container_width=True, hide_index=True)
        
        with tab4:
            st.markdown("#### 🔥 Mapa de Calor: Concentración de Personal")
            top_deps_heat = datos_filtrados[dependencia_col].value_counts().head(15).index
            top_jer_heat = datos_filtrados[jerarquia_col].value_counts().head(10).index
            
            datos_heat = datos_filtrados[datos_filtrados[dependencia_col].isin(top_deps_heat)]
            datos_heat = datos_heat[datos_heat[jerarquia_col].isin(top_jer_heat)]
            
            if len(datos_heat) > 0:
                heatmap_data = pd.crosstab(
                    datos_heat[dependencia_col], 
                    datos_heat[jerarquia_col]
                ).fillna(0)
                
                if len(heatmap_data) > 0 and len(heatmap_data.columns) > 0:
                    fig_heatmap = px.imshow(
                        heatmap_data,
                        text_auto=True,
                        aspect="auto",
                        color_continuous_scale='Reds',
                        title="Concentración de personal",
                        labels=dict(x="Jerarquía", y="Dependencia", color="Cantidad")
                    )
                    fig_heatmap.update_layout(height=600)
                    st.plotly_chart(fig_heatmap, use_container_width=True)
            
            st.markdown("#### 💡 Gráfico de Burbujas")
            bubble_data = datos_filtrados.groupby(dependencia_col).agg({
                jerarquia_col: 'nunique',
                identificador_col: 'count' if identificador_col else 'size'
            }).reset_index()
            bubble_data.columns = ['Dependencia', 'Diversidad Jerarquías', 'Total Personal']
            
            if len(bubble_data) > 0:
                fig_bubble = px.scatter(
                    bubble_data,
                    x='Total Personal',
                    y='Diversidad Jerarquías',
                    size='Total Personal',
                    color='Dependencia',
                    hover_name='Dependencia',
                    title="Relación: Personal total vs Diversidad de jerarquías"
                )
                fig_bubble.update_layout(height=500)
                st.plotly_chart(fig_bubble, use_container_width=True)
        
        # ========== RESUMEN TABLA (solo admin y supervisor) ==========
        st.markdown("## 📊 Resumen por Jerarquía y Dependencia")
        if len(datos_filtrados) > 0:
            datos_filtrados = datos_filtrados.reset_index(drop=True)
            resumen_df = datos_filtrados.groupby([dependencia_col, jerarquia_col]).size().reset_index(name='Cantidad')
            resumen_df['Cantidad'] = pd.to_numeric(resumen_df['Cantidad'], errors='coerce').fillna(0).astype(int)
            resumen = resumen_df.pivot(index=dependencia_col, columns=jerarquia_col, values='Cantidad').fillna(0)
            for col in resumen.columns:
                resumen[col] = pd.to_numeric(resumen[col], errors='coerce').fillna(0).astype(int)
            resumen['Total'] = resumen.sum(axis=1).astype(int)
            total_row = resumen.sum(axis=0).to_frame().T
            total_row.index = ['Total']
            total_row = total_row.astype(int)
            resumen = pd.concat([resumen, total_row])
            
            columnas_jerarquia = [col for col in orden_jerarquia_base if col in resumen.columns]
            otras_columnas = [col for col in resumen.columns if col not in columnas_jerarquia and col != 'Total']
            columnas_ordenadas = columnas_jerarquia + sorted(otras_columnas) + ['Total']
            resumen = resumen[columnas_ordenadas]
            
            df_mostrar = resumen.reset_index()
            primera_col = df_mostrar.columns[0]
            df_mostrar = df_mostrar.rename(columns={primera_col: 'Dependencia'})
            df_mostrar = df_mostrar.loc[:, ~df_mostrar.columns.duplicated()]
            
            st.dataframe(df_mostrar, use_container_width=True, hide_index=True)
        else:
            st.info("ℹ️ No hay datos para generar el resumen")
    
    # ========== LISTADO DE PERSONAL POR DEPENDENCIA ==========
    st.markdown("## 📋 Listado de Personal")
    st.caption(f"Total de registros: {len(datos_filtrados)}")
    
    if len(datos_filtrados) == 0:
        st.warning("⚠️ No hay datos con los filtros seleccionados")
    else:
        # Agrupar por dependencia
        for dependencia, grupo in datos_filtrados.groupby(dependencia_col):
            # Título de la dependencia con estilo
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #f8fafc 0%, #e8edf4 100%); 
                        padding: 12px 20px; 
                        border-radius: 10px; 
                        border-left: 4px solid #2ecc71;
                        margin: 15px 0 10px 0;">
                <h3 style="margin: 0; color: #1f3a6b;">🏢 {dependencia}</h3>
            </div>
            """, unsafe_allow_html=True)
            
            # Preparar los datos del grupo (sin ID)
            df_grupo = grupo[[nombre_col, jerarquia_col, funcion_col]].copy()
            df_grupo = df_grupo.rename(columns={
                nombre_col: 'Nombre',
                jerarquia_col: 'Jerarquía',
                funcion_col: 'Función'
            })
            
            # Mostrar tabla con selección de fila
            event = st.dataframe(
                df_grupo,
                use_container_width=True,
                hide_index=True,
                selection_mode="single-row",
                on_select="rerun"
            )
            
            # Verificar si se seleccionó una fila
            if event and event.selection and "rows" in event.selection and len(event.selection["rows"]) > 0:
                row_idx = event.selection["rows"][0]
                if row_idx is not None:
                    idx_original = grupo.index[row_idx]
                    agente = datos_filtrados.loc[idx_original]
                    
                    # ========== FICHA DEL AGENTE EN FORMATO TARJETA ==========
                    st.markdown("---")
                    
                    # Inicializar variables para la ficha
                    nombre_agente = agente.get(nombre_col, 'Sin nombre')
                    jerarquia_agente = agente.get(jerarquia_col, 'Sin jerarquía')
                    funcion_agente = agente.get(funcion_col, 'Sin función')
                    dependencia_agente = agente.get(dependencia_col, 'Sin dependencia')
                    
                    # Obtener iniciales para el avatar
                    iniciales = ''.join([p[0] for p in nombre_agente.split()[:2]]) if nombre_agente else '?'
                    
                    # Construir la ficha HTML
                    ficha_html = f"""
                    <div class="agente-ficha">
                        <div class="agente-header">
                            <div class="agente-avatar">{iniciales.upper()}</div>
                            <div>
                                <div class="agente-nombre">{nombre_agente}</div>
                                <div style="display: flex; gap: 8px; flex-wrap: wrap; margin-top: 6px;">
                                    <span class="agente-badge badge-jerarquia">⭐ {jerarquia_agente}</span>
                                    <span class="agente-badge badge-funcion">📋 {funcion_agente}</span>
                                    <span class="agente-badge badge-dependencia">🏢 {dependencia_agente}</span>
                                </div>
                            </div>
                        </div>
                        <div class="agente-grid">
                    """
                    
                    # Agregar todos los campos del agente (excepto los ya mostrados)
                    campos_excluidos = [nombre_col, jerarquia_col, funcion_col, dependencia_col]
                    for col in agente.index:
                        if col in campos_excluidos:
                            continue
                        valor = agente[col]
                        # Limpiar valor NaN
                        if pd.isna(valor) or str(valor).strip() == '' or str(valor).strip() == 'nan':
                            valor_display = '<span class="agente-campo-valor vacio">(sin dato)</span>'
                        else:
                            valor_display = f'<span class="agente-campo-valor">{valor}</span>'
                        
                        # Limpiar nombre de columna para display
                        nombre_display = col.replace('_', ' ').title()
                        ficha_html += f"""
                            <div class="agente-campo">
                                <div class="agente-campo-label">{nombre_display}</div>
                                {valor_display}
                            </div>
                        """
                    
                    ficha_html += """
                        </div>
                    </div>
                    """
                    
                    st.markdown(ficha_html, unsafe_allow_html=True)
            
            st.markdown("---")
    
    # ========== MÓDULO DE ROTACIÓN (SOLO ADMIN) ==========
    if es_admin:
        st.markdown("---")
        st.markdown("## 🔄 Rotación de Personal")
        with st.expander("⚙️ Configurar rotación", expanded=False):
            dependencias_totales = sorted(datos_completos[dependencia_col].unique())
            deps_rotacion = st.multiselect("Seleccione dependencias que participarán:", dependencias_totales)
            jerarquias_rotacion = st.multiselect("Seleccione jerarquías a rotar:", opciones_jerarquia)
            
            if deps_rotacion and jerarquias_rotacion:
                candidatos_df = datos_completos[datos_completos[dependencia_col].isin(deps_rotacion)]
                candidatos_df = candidatos_df[candidatos_df[jerarquia_col].isin(jerarquias_rotacion)]
                if len(candidatos_df) == 0:
                    st.warning("⚠️ No hay personal en las dependencias y jerarquías seleccionadas.")
                else:
                    st.markdown("#### ✏️ Seleccione agentes a rotar")
                    st.caption("Desmarque los agentes que NO deben rotar")
                    columnas_a_mostrar = ['index', identificador_col, nombre_col, jerarquia_col, funcion_col, dependencia_col]
                    if sexo_col:
                        columnas_a_mostrar.append(sexo_col)
                    candidatos_con_idx = candidatos_df.reset_index()[columnas_a_mostrar]
                    candidatos_con_idx.insert(1, 'Rotar', True)
                    with st.form(key="rotacion_form"):
                        edited_df = st.data_editor(
                            candidatos_con_idx,
                            column_config={
                                "Rotar": st.column_config.CheckboxColumn("🔄 Rotar", default=True),
                                "index": None
                            },
                            hide_index=True,
                            use_container_width=True,
                            height=400
                        )
                        submitted = st.form_submit_button("🎲 Generar propuesta", use_container_width=True, type="primary")
                    
                    if submitted:
                        seleccionados = edited_df[edited_df['Rotar'] == True]
                        if len(seleccionados) == 0:
                            st.error("❌ No ha seleccionado ningún agente para rotar")
                        else:
                            indices_seleccionados = seleccionados['index'].tolist()
                            seleccionados_df = datos_completos.loc[indices_seleccionados].copy()
                            
                            def rotar_grupo(grupo_df, deps):
                                if len(grupo_df) == 0:
                                    return grupo_df
                                conteo_origen = grupo_df.groupby(dependencia_col).size().to_dict()
                                shuffled = grupo_df.sample(frac=1, random_state=random.randint(1, 10000)).copy()
                                nuevas_asignaciones = []
                                for dep, count in conteo_origen.items():
                                    nuevas_asignaciones.extend([dep] * count)
                                random.shuffle(nuevas_asignaciones)
                                shuffled['DEPENDENCIA_DESTINO'] = nuevas_asignaciones
                                return shuffled
                            
                            def agregar_intercambio_unico(propuesta_df):
                                propuesta_df = propuesta_df[propuesta_df['Origen'] != propuesta_df['Destino']].copy()
                                if propuesta_df.empty:
                                    return propuesta_df
                                mapping = defaultdict(list)
                                for _, row in propuesta_df.iterrows():
                                    key = (row['Origen'], row['Destino'])
                                    mapping[key].append(row['Agente'])
                                intercambios = []
                                pares_procesados = set()
                                filas_a_conservar = []
                                for idx, row in propuesta_df.iterrows():
                                    origen = row['Origen']
                                    destino = row['Destino']
                                    par = tuple(sorted([origen, destino]))
                                    if par in pares_procesados:
                                        continue
                                    clave_inversa = (destino, origen)
                                    candidatos = mapping.get(clave_inversa, [])
                                    if candidatos:
                                        elegido = candidatos[0]
                                        intercambios.append(elegido)
                                    else:
                                        intercambios.append("(sin intercambio directo)")
                                    filas_a_conservar.append(idx)
                                    pares_procesados.add(par)
                                propuesta_filtrada = propuesta_df.loc[filas_a_conservar].copy()
                                propuesta_filtrada['Intercambia con'] = intercambios
                                return propuesta_filtrada
                            
                            todas_propuestas = []
                            for jerarquia in jerarquias_rotacion:
                                jer_df = seleccionados_df[seleccionados_df[jerarquia_col] == jerarquia]
                                if len(jer_df) == 0:
                                    continue
                                if sexo_col and sexo_col in jer_df.columns:
                                    sexos = jer_df[sexo_col].unique()
                                    grupos_rotados = []
                                    for s in sexos:
                                        grupo = jer_df[jer_df[sexo_col] == s]
                                        if len(grupo) > 0:
                                            grupos_rotados.append(rotar_grupo(grupo, deps_rotacion))
                                    shuffled_total = pd.concat(grupos_rotados, ignore_index=True) if grupos_rotados else jer_df
                                else:
                                    shuffled_total = rotar_grupo(jer_df, deps_rotacion)
                                columnas_propuesta = [identificador_col, nombre_col, jerarquia_col, funcion_col, dependencia_col, 'DEPENDENCIA_DESTINO']
                                if sexo_col and sexo_col in shuffled_total.columns:
                                    columnas_propuesta.append(sexo_col)
                                prop = shuffled_total[columnas_propuesta].copy()
                                prop.rename(columns={
                                    nombre_col: 'Agente',
                                    dependencia_col: 'Origen',
                                    'DEPENDENCIA_DESTINO': 'Destino',
                                    identificador_col: 'Identificador',
                                    jerarquia_col: 'Jerarquía',
                                    funcion_col: 'Función'
                                }, inplace=True)
                                if sexo_col and sexo_col in shuffled_total.columns:
                                    prop.rename(columns={sexo_col: 'Sexo'}, inplace=True)
                                prop = agregar_intercambio_unico(prop)
                                if not prop.empty:
                                    todas_propuestas.append(prop)
                            if todas_propuestas:
                                propuesta_final = pd.concat(todas_propuestas, ignore_index=True)
                                if propuesta_final.empty:
                                    st.warning("⚠️ Todos los agentes terminan en la misma dependencia")
                                else:
                                    st.session_state.propuesta_rotacion = propuesta_final
                                    st.success("✅ Propuesta generada exitosamente")
                                    st.rerun()
                            else:
                                st.warning("⚠️ No se generaron cambios reales")
            else:
                if not deps_rotacion:
                    st.info("ℹ️ Seleccione al menos una dependencia")
                elif not jerarquias_rotacion:
                    st.info("ℹ️ Seleccione al menos una jerarquía")
        
        if st.session_state.propuesta_rotacion is not None:
            st.markdown("## 📋 Propuesta de intercambio")
            st.dataframe(st.session_state.propuesta_rotacion, use_container_width=True, hide_index=True)
            col_conf1, col_conf2 = st.columns(2)
            with col_conf1:
                if st.button("✅ Confirmar rotación", use_container_width=True, type="primary"):
                    try:
                        sh = get_new_connection()
                        ws = sh.worksheet("Personal")
                        all_data = ws.get_all_values()
                        header = all_data[0]
                        col_dep_idx = header.index(dependencia_col)
                        col_id_idx = header.index(identificador_col)
                        col_nombre_idx = header.index(nombre_col)
                        with st.spinner("Aplicando rotación..."):
                            cambios_detalle = []
                            for _, row in st.session_state.propuesta_rotacion.iterrows():
                                for i, data_row in enumerate(all_data[1:], start=2):
                                    if data_row[col_id_idx] == row['Identificador'] and data_row[col_nombre_idx] == row['Agente']:
                                        origen = data_row[col_dep_idx]
                                        destino = row['Destino']
                                        if origen != destino:
                                            ws.update_cell(i, col_dep_idx+1, destino)
                                            cambios_detalle.append(f"{row['Agente']}: {origen} → {destino}")
                                        break
                            # Registrar en auditoría
                            if cambios_detalle:
                                detalle = "Rotación: " + "; ".join(cambios_detalle)
                                registrar_auditoria(user['DNI'], user['NOMBRE'], "Varias", "ROTACION", detalle, sh)
                            
                            reordenar_por_jerarquia()
                            resultado = cargar_datos_hoja()
                            st.session_state.df_personal = resultado[0]
                            st.session_state.df_usuarios = resultado[1]
                            st.session_state.df_propuestas = resultado[2]
                            st.session_state.df_auditoria = resultado[3]
                            st.session_state.sh = resultado[4]
                            st.session_state.propuesta_rotacion = None
                            st.success("✅ Rotación aplicada correctamente")
                            time.sleep(1)
                            st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error al actualizar: {e}")
            with col_conf2:
                if st.button("❌ Cancelar propuesta", use_container_width=True):
                    st.session_state.propuesta_rotacion = None
                    st.rerun()
    
    # ========== EXPORTAR (para todos) ==========
    st.markdown("---")
    st.markdown("## 📎 Exportar datos")
    col_export1, col_export2 = st.columns(2)
    with col_export1:
        if OPENPYXL_AVAILABLE:
            formato = st.radio("Formato:", ["CSV", "Excel (XLSX)"], horizontal=True)
        else:
            formato = "CSV"
            st.caption("💡 Para exportar a Excel, instale openpyxl")
    with col_export2:
        if st.button("📥 Exportar listado", use_container_width=True):
            if len(datos_filtrados) == 0:
                st.warning("⚠️ No hay datos para exportar")
            else:
                nombre_base = f"personal_{datetime.datetime.now(tz_arg).strftime('%Y%m%d_%H%M%S')}"
                if formato == "CSV":
                    csv = datos_filtrados.to_csv(index=False).encode('utf-8-sig')
                    st.download_button("✅ Descargar CSV", csv, f"{nombre_base}.csv", mime="text/csv")
                elif OPENPYXL_AVAILABLE:
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        datos_filtrados.to_excel(writer, sheet_name='Personal', index=False)
                        workbook = writer.book
                        hoja = writer.sheets['Personal']
                        for col in hoja.columns:
                            max_len = 0
                            col_letter = get_column_letter(col[0].column)
                            for cell in col:
                                try:
                                    max_len = max(max_len, len(str(cell.value)))
                                except:
                                    pass
                            hoja.column_dimensions[col_letter].width = min(max_len + 2, 50)
                        for celda in hoja[1]:
                            celda.font = Font(bold=True)
                            celda.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                            celda.alignment = Alignment(horizontal='center')
                        for row in hoja.iter_rows(min_row=2):
                            for cell in row:
                                if isinstance(cell.value, (int, float)):
                                    cell.alignment = Alignment(horizontal='center')
                        hoja.freeze_panes = 'A2'
                    st.download_button("✅ Descargar Excel", output.getvalue(), f"{nombre_base}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
    st.markdown(
        "<div class='footer'>"
        "© 2024 - Sistema de Gestión de Personal | Desarrollado para la Fuerza Policial | Versión 3.0"
        "</div>", 
        unsafe_allow_html=True
    )
